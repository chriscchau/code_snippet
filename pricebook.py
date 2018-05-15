#! /usr/bin/python3
#

import openpyxl, pprint
from openpyxl.styles import Font
print ('Opening workbook...')
wb = openpyxl.load_workbook('HK_26042018_FY18_1Q_nopw.xlsx')
sheetNames = wb.get_sheet_names()
priceData = {}

priceBook = openpyxl.Workbook()
serverProducts = priceBook.get_active_sheet()
serverProducts.title = 'Server Products'

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False

# Fill in priceData
print('Reading rows...')

row_index = 0
for sheet in sheetNames:
    if sheet == '100 (h Series)' or sheet == '100' or sheet == 'ft' or sheet == 'ftsys' or sheet == 'Mona' or sheet == 'Tape' or sheet == 'MEMDump':
        sh = wb.get_sheet_by_name(sheet)
        row_count = sh.max_row
        row_index = row_index + 5
        row_write = row_index

        priceData.clear()
        for row in range(5, row_count + 1):
            row_index = row_index + 1

            n_code = sh['A' + str(row)].value
            product = sh['B' + str(row)].value
            tp = sh['D' + str(row)].value

            if tp is not None and is_number(tp) is True:
                NECHK_Cost = tp * 7.8

                if sheet == 'ft' or sheet == 'ftsys':
                    NECHK_LP = NECHK_Cost * 3
                else:
                    NECHK_LP = NECHK_Cost * 2.5

                resellerCost = NECHK_LP * 0.9
                registeredCost = NECHK_LP * 0.65
                registeredCost1 = NECHK_LP * 0.55
                registeredCost2 = NECHK_LP * 0.5

                priceData[row] = [n_code, product, tp, NECHK_Cost, NECHK_LP, resellerCost, registeredCost, registeredCost1, registeredCost2]

        print(sheet)
        print(row_index)

        serverProducts['A' + str(row_write - 4)] = sheet
        serverProducts['A' + str(row_write - 4)].font = Font(bold=True, size=16)
        serverProducts['B' + str(row_write - 1)] = 'N-Code'
        serverProducts['C' + str(row_write - 1)] = 'Product'
        serverProducts['D' + str(row_write - 1)] = 'TP (USD)'
        serverProducts['E' + str(row_write - 1)] = 'TP (HKD)'
        serverProducts['F' + str(row_write - 1)] = 'List Price (HKD)'
        serverProducts['G' + str(row_write - 1)] = 'Reseller Cost (HKD)'
        serverProducts['H' + str(row_write - 1)] = 'Registered Cost (HKD)'
        serverProducts['I' + str(row_write - 1)] = '1st Discount (HKD)'
        serverProducts['J' + str(row_write - 1)] = '2nd Discount (HKD)'

        serverProducts['B' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['C' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['D' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['E' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['F' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['G' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['H' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['I' + str(row_write - 1)].font = Font(bold=True)
        serverProducts['J' + str(row_write - 1)].font = Font(bold=True)

        for newRow in range(5, row_count + 1):

            serverProducts['B' + str(row_write)] = priceData[newRow][0]
            serverProducts['C' + str(row_write)] = priceData[newRow][1]
            serverProducts['D' + str(row_write)] = priceData[newRow][2]
            serverProducts['E' + str(row_write)] = priceData[newRow][3]
            serverProducts['F' + str(row_write)] = priceData[newRow][4]
            serverProducts['G' + str(row_write)] = priceData[newRow][5]
            serverProducts['H' + str(row_write)] = priceData[newRow][6]
            serverProducts['I' + str(row_write)] = priceData[newRow][7]
            serverProducts['J' + str(row_write)] = priceData[newRow][8]

            serverProducts['B' + str(row_write)].font = Font(size=10)
            serverProducts['C' + str(row_write)].font = Font(size=10)
            serverProducts['D' + str(row_write)].font = Font(size=10)
            serverProducts['E' + str(row_write)].font = Font(size=10)
            serverProducts['F' + str(row_write)].font = Font(size=10)
            serverProducts['G' + str(row_write)].font = Font(size=10)
            serverProducts['H' + str(row_write)].font = Font(size=10)
            serverProducts['I' + str(row_write)].font = Font(size=10)
            serverProducts['J' + str(row_write)].font = Font(size=10)

            serverProducts['D' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'
            serverProducts['E' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'
            serverProducts['F' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'
            serverProducts['G' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'
            serverProducts['H' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'
            serverProducts['I' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'
            serverProducts['J' + str(row_write)].number_format = '"$"* #,##0.00_);("$"* #,##0.00)'

            row_write = row_write + 1


sheetNames = wb.get_sheet_names()
for name in sheetNames:
    print(name)

priceBook.save('pricebook.xlsx')
